"""
تطبيق Streamlit لعرض نتائج الطلاب
يدعم المرحلتين الثانية والثالثة:
- الأكواد التي تبدأ بـ 2 (مثل 2001) → المرحلة الثانية → m2.xlsx + pay2.xlsx
- الأكواد التي تبدأ بـ 3 (مثل 3001) → المرحلة الثالثة → m3.xlsx + pay3.xlsx
"""

import streamlit as st
from openpyxl import load_workbook
import re
import os

# ===================== إعدادات الصفحة =====================
st.set_page_config(
    page_title="نظام عرض نتائج الطلاب",
    page_icon="🎓",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# ===================== CSS لدعم اللغة العربية =====================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@400;500;700;800&display=swap');

    /* ===== RTL شامل لكل عناصر التطبيق ===== */
    html, body, [class*="css"], .stApp, .main, .block-container,
    div, span, p, h1, h2, h3, h4, h5, h6, label, button, input, textarea, select,
    [data-testid="stAppViewContainer"],
    [data-testid="stHeader"],
    [data-testid="stSidebar"],
    [data-testid="stMarkdown"],
    [data-testid="stMarkdownContainer"],
    [data-testid="stText"],
    [data-testid="stAlert"],
    [data-testid="stNotification"],
    [data-testid="stWarning"],
    [data-testid="stError"],
    [data-testid="stSuccess"],
    [data-testid="stInfo"],
    [data-testid="stTextInput"],
    [data-testid="stButton"],
    [data-testid="stForm"],
    [data-testid="column"] {
        direction: rtl !important;
        text-align: right !important;
    }

    /* الخط العربي */
    * {
        font-family: 'Tajawal', sans-serif !important;
    }

    /* خلفية التطبيق */
    .stApp {
        background: linear-gradient(135deg, #f5f7fa 0%, #e4ecf7 100%);
    }

    /* العناوين دائماً وسط */
    h1, h2, h3, h4, h5, h6 {
        text-align: center !important;
        color: #1a365d;
        font-weight: 800 !important;
    }

    /* ===== ترويسة الموقع ===== */
    .header-box {
        background: linear-gradient(135deg, #1a365d 0%, #2d5a8e 100%);
        color: white;
        padding: 2rem 1.5rem;
        border-radius: 15px;
        text-align: center !important;
        margin-bottom: 2rem;
        box-shadow: 0 8px 24px rgba(26, 54, 93, 0.2);
        direction: rtl;
    }

    .header-box h1 {
        color: white !important;
        margin: 0;
        font-size: 2rem;
        text-align: center !important;
    }

    .header-box p {
        margin: 0.5rem 0 0 0;
        opacity: 0.9;
        font-size: 1.1rem;
        text-align: center !important;
    }

    .stage-badge {
        display: inline-block;
        background: rgba(255,255,255,0.2);
        padding: 0.3rem 1rem;
        border-radius: 20px;
        font-size: 0.95rem;
        font-weight: 700;
        margin-top: 0.5rem;
    }

    /* ===== صناديق المعلومات ===== */
    .student-info {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        border-right: 5px solid #2d5a8e;
        margin: 1rem 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        direction: rtl;
        text-align: right;
    }

    .student-info h3 {
        color: #1a365d;
        margin-top: 0;
        text-align: right !important;
    }

    .student-info p {
        text-align: right !important;
    }

    /* ===== بطاقات الدرجات ===== */
    .grade-card {
        background: white;
        padding: 1.2rem;
        border-radius: 10px;
        margin: 0.8rem 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        border-right: 4px solid #38a169;
        direction: rtl;
    }

    .grade-card-missing {
        background: #fff5f5;
        padding: 1.2rem;
        border-radius: 10px;
        margin: 0.8rem 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        border-right: 4px solid #e53e3e;
        direction: rtl;
    }

    .subject-name {
        font-size: 1.2rem;
        font-weight: 700;
        color: #1a365d;
        margin-bottom: 0.7rem;
        text-align: right;
    }

    .grade-row {
        display: flex;
        flex-direction: row-reverse;
        justify-content: space-between;
        align-items: center;
        padding: 0.4rem 0;
    }

    .grade-label { color: #4a5568; font-weight: 500; }

    .grade-value {
        font-weight: 700;
        font-size: 1.1rem;
        color: #2d5a8e;
        background: #ebf4ff;
        padding: 0.3rem 0.8rem;
        border-radius: 6px;
        direction: ltr;  /* الأرقام بشكل LTR لقراءة صحيحة */
    }

    .grade-missing { color: #c53030; font-style: italic; }

    /* ===== حقل الإدخال ===== */
    .stTextInput > div > div > input {
        text-align: center !important;
        font-size: 1.3rem;
        font-weight: 700;
        padding: 0.7rem;
        border-radius: 10px;
        border: 2px solid #2d5a8e;
        direction: ltr;  /* الأرقام تكتب LTR */
    }

    /* ===== الأزرار ===== */
    .stButton > button {
        width: 100%;
        background: linear-gradient(135deg, #2d5a8e 0%, #1a365d 100%);
        color: white;
        font-size: 1.1rem;
        font-weight: 700;
        padding: 0.7rem;
        border: none;
        border-radius: 10px;
        margin-top: 0.5rem;
    }

    .stButton > button:hover {
        background: linear-gradient(135deg, #1a365d 0%, #0d2240 100%);
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(26, 54, 93, 0.3);
    }

    /* ===== رسائل التنبيه (alerts) ===== */
    [data-testid="stAlert"], [data-testid="stNotification"],
    div[role="alert"] {
        direction: rtl !important;
        text-align: right !important;
    }

    [data-testid="stAlert"] *, [data-testid="stNotification"] *,
    div[role="alert"] * {
        text-align: right !important;
    }

    /* أيقونة التنبيه على اليمين */
    [data-testid="stAlert"] > div:first-child,
    [data-testid="stNotification"] > div:first-child {
        order: 1;
    }

    /* ===== التذييل ===== */
    .footer {
        text-align: center !important;
        color: #718096;
        margin-top: 3rem;
        padding-top: 1rem;
        border-top: 1px solid #e2e8f0;
        font-size: 0.9rem;
        line-height: 1.8;
        direction: rtl;
    }

    .footer .developer {
        color: #2d5a8e;
        font-weight: 700;
        font-size: 0.95rem;
    }

    /* إخفاء عناصر Streamlit الافتراضية غير المرغوبة */
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }

    /* تأكيد الاتجاه على المستوى العام */
    body {
        direction: rtl;
    }
</style>
""", unsafe_allow_html=True)


# ===================== إعدادات المراحل =====================
# يمكن إضافة مراحل جديدة هنا بسهولة (مثلاً: '4': {...} للمرحلة الرابعة)
STAGES = {
    '2': {
        'name': 'المرحلة الثانية',
        'grades_file': 'm2.xlsx',
        'payments_file': 'pay2.xlsx'
    },
    '3': {
        'name': 'المرحلة الثالثة',
        'grades_file': 'm3.xlsx',
        'payments_file': 'pay3.xlsx'
    },
}


# ===================== دوال مساعدة =====================
def normalize_arabic_name(name):
    """تنظيف وتوحيد الأسماء العربية للمطابقة"""
    if name is None:
        return ""
    name = str(name).strip()

    # إزالة التشكيل
    arabic_diacritics = re.compile(r'[\u064B-\u065F\u0670]')
    name = arabic_diacritics.sub('', name)

    # توحيد الحروف المتشابهة
    name = name.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا')
    name = name.replace('ى', 'ي').replace('ئ', 'ي').replace('ؤ', 'و')
    name = name.replace('ة', 'ه')

    # توحيد المسافات (إزالة المسافات الزائدة)
    name = re.sub(r'\s+', ' ', name).strip()

    # توحيد الأسماء المركبة:
    # المشكلة: قد يُكتب "عبد العزيز" بمسافة أو "عبدالعزيز" بدون مسافة
    # الحل: إزالة المسافة بين "عبد" والكلمة التالية إن بدأت بـ "ال"
    # نفس المشكلة للأسماء المركبة الأخرى مثل: ابو، ام، ذو
    compound_prefixes = ['عبد', 'ابو', 'ام', 'ابن', 'ذو', 'ذي', 'ذا']
    for prefix in compound_prefixes:
        # نمط: <prefix> ال<شيء> → <prefix>ال<شيء>
        # مثال: "عبد العزيز" → "عبدالعزيز"، "ابو الحسن" → "ابوالحسن"
        pattern = r'\b' + prefix + r'\s+ال'
        name = re.sub(pattern, prefix + 'ال', name)

    return name


def find_payment_match(student_name, payments_dict, min_parts=3):
    """
    البحث عن المبلغ المتبقي للطالب في قاموس الأقساط.

    خطوات المطابقة بالترتيب:
    1️⃣ مطابقة كاملة (الاسم بالكامل)
    2️⃣ مطابقة جزئية: اسم الطالب من ملف الدرجات يكون **بداية** اسم في ملف المالية
       (مثلاً: 'ابراهيم عبد الكريم فليح سلمان' يطابق 'ابراهيم عبد الكريم فليح سلمان العماري')
       - يشترط أن يحوي اسم الطالب على الأقل min_parts كلمات (افتراضياً 3) لتجنب المطابقات الخاطئة
       - يشترط أن تكون مطابقة فريدة (طالب واحد فقط) في ملف المالية
    3️⃣ المطابقة العكسية: اسم في ملف المالية يكون بداية اسم الطالب
       (في حالة أن ملف المالية يحوي اسماً أقصر)

    Returns:
        (remaining_amount, match_type) حيث match_type يكون: 'exact', 'partial', 'reverse', أو None
    """
    student_name = student_name.strip()
    if not student_name:
        return None, None

    # 1️⃣ مطابقة كاملة
    if student_name in payments_dict:
        return payments_dict[student_name], 'exact'

    # تجزئة الاسم لعدد الكلمات
    student_parts = student_name.split()
    if len(student_parts) < min_parts:
        return None, None  # الاسم قصير جداً، لا نخاطر بمطابقة جزئية

    # 2️⃣ مطابقة جزئية: اسم الطالب بداية لاسم في ملف المالية
    # نبحث عن جميع الأسماء التي تبدأ بنفس الاسم الكامل للطالب + مسافة
    student_prefix = student_name + ' '
    candidates_forward = [
        pay_name for pay_name in payments_dict.keys()
        if pay_name.startswith(student_prefix)
    ]

    if len(candidates_forward) == 1:
        # مطابقة فريدة - ممتاز
        return payments_dict[candidates_forward[0]], 'partial'
    elif len(candidates_forward) > 1:
        # تعدد - لا نخاطر
        return None, 'ambiguous'

    # 3️⃣ المطابقة العكسية: اسم في ملف المالية يكون بداية لاسم الطالب
    # (مثلاً ملف الدرجات: 'احمد علي حسن العماري' وملف المالية: 'احمد علي حسن')
    candidates_reverse = []
    for pay_name in payments_dict.keys():
        pay_parts = pay_name.split()
        if len(pay_parts) < min_parts:
            continue
        if student_name.startswith(pay_name + ' '):
            candidates_reverse.append(pay_name)

    if len(candidates_reverse) == 1:
        return payments_dict[candidates_reverse[0]], 'reverse'

    return None, None


def parse_amount(value):
    """تحويل المبلغ من نص (قد يحوي فواصل) إلى رقم"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace(',', '').replace('،', '').strip()
    if s == '':
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_grade(value):
    """تحويل الدرجة - يعيد None إذا كانت فارغة"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s == '':
        return None
    try:
        return float(s)
    except ValueError:
        return None


def detect_stage(code):
    """يحدد المرحلة من أول رقم في الكود"""
    code = str(code).strip()
    if not code:
        return None
    first_char = code[0]
    return first_char if first_char in STAGES else None


# ===================== تحميل البيانات =====================
@st.cache_data
def load_payments(file_path):
    """قراءة ملف الأقساط وبناء قاموس: اسم_مطبّع -> مبلغ متبقي"""
    if not os.path.exists(file_path):
        return None, f"ملف الأقساط '{file_path}' غير موجود"

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    name_col = None
    remaining_col = None

    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            v = str(val).strip()
            if 'اسم' in v and 'طالب' in v:
                name_col = col_idx
            elif 'متبقي' in v:
                remaining_col = col_idx

    if name_col is None or remaining_col is None:
        return None, f"لم يتم العثور على أعمدة 'اسم الطالب' أو 'المبلغ المتبقي' في '{file_path}'"

    payments = {}
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=name_col).value
        remaining = ws.cell(row=row_idx, column=remaining_col).value
        if name:
            normalized = normalize_arabic_name(name)
            if normalized:
                payments[normalized] = parse_amount(remaining)

    return payments, None


@st.cache_data
def load_grades(file_path):
    """
    قراءة ملف الدرجات.
    البنية:
    - الصف 1: ت | الطالب | اسم المادة 1 (مدمج) | اسم المادة 2 (مدمج) | ...
    - الصف 2: (فارغ) | (فارغ) | السعي (40%) | المد (10%) | السعي (40%) | المد (10%) | ...
    - الصف 3+: بيانات الطلاب
    """
    if not os.path.exists(file_path):
        return None, f"ملف الدرجات '{file_path}' غير موجود"

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    code_col = None
    name_col = None

    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            v = str(val).strip()
            if v == 'ت' or v == 'الكود' or v == '#':
                code_col = col_idx
            elif 'طالب' in v:
                name_col = col_idx

    if code_col is None:
        return None, f"لم يتم العثور على عمود الكود (ت) في '{file_path}'"
    if name_col is None:
        return None, f"لم يتم العثور على عمود 'الطالب' في '{file_path}'"

    # اكتشاف المواد: اسم المادة في الصف 1، والصف 2 يحوي "السعي" و"المد"
    subjects = []
    col_idx = 1
    while col_idx <= ws.max_column:
        val_row1 = ws.cell(row=1, column=col_idx).value
        if val_row1:
            v = str(val_row1).strip()
            if v in ['ت', 'الطالب', '#', 'الكود'] or 'طالب' in v:
                col_idx += 1
                continue

            sub1 = ws.cell(row=2, column=col_idx).value
            sub2 = ws.cell(row=2, column=col_idx + 1).value if col_idx + 1 <= ws.max_column else None

            sub1_str = str(sub1).strip() if sub1 else ''
            sub2_str = str(sub2).strip() if sub2 else ''

            if 'سعي' in sub1_str and 'مد' in sub2_str:
                subjects.append({
                    'name': v,
                    'sa3y_col': col_idx,
                    'mid_col': col_idx + 1
                })
                col_idx += 2
                continue
        col_idx += 1

    if not subjects:
        return None, f"لم يتم العثور على أي مواد في '{file_path}'"

    students = {}
    for row_idx in range(3, ws.max_row + 1):
        code = ws.cell(row=row_idx, column=code_col).value
        name = ws.cell(row=row_idx, column=name_col).value

        if code is None or name is None:
            continue

        code_str = str(code).strip()
        if code_str.endswith('.0'):
            code_str = code_str[:-2]
        if not code_str:
            continue

        name_str = str(name).strip()
        if not name_str:
            continue

        student_subjects = {}
        for subj in subjects:
            sa3y = parse_grade(ws.cell(row=row_idx, column=subj['sa3y_col']).value)
            mid = parse_grade(ws.cell(row=row_idx, column=subj['mid_col']).value)
            student_subjects[subj['name']] = {'sa3y': sa3y, 'mid': mid}

        students[code_str] = {
            'name': name_str,
            'normalized_name': normalize_arabic_name(name_str),
            'subjects': student_subjects
        }

    return students, None


@st.cache_data
def load_all_stages():
    """تحميل بيانات جميع المراحل دفعة واحدة"""
    all_data = {}
    errors = []

    for stage_key, stage_info in STAGES.items():
        students, err1 = load_grades(stage_info['grades_file'])
        payments, err2 = load_payments(stage_info['payments_file'])

        if err1:
            errors.append(f"⚠️ {stage_info['name']}: {err1}")
            continue
        if err2:
            errors.append(f"⚠️ {stage_info['name']}: {err2}")
            continue

        all_data[stage_key] = {
            'name': stage_info['name'],
            'students': students,
            'payments': payments
        }

    return all_data, errors


# ===================== عرض النتائج =====================
def display_student_results(student_data, stage_name):
    """عرض نتائج الطالب بعد التحقق من تسديد القسط"""
    st.markdown(f"""
    <div class="student-info">
        <h3>📋 معلومات الطالب</h3>
        <p style="font-size: 1.2rem; margin: 0.5rem 0;">
            <strong>الاسم:</strong> {student_data['name']}
        </p>
        <p style="font-size: 1rem; margin: 0.5rem 0; color: #2d5a8e;">
            <strong>المرحلة:</strong> {stage_name}
        </p>
        <p style="font-size: 1rem; margin: 0; color: #38a169;">
            <strong>✅ القسط مسدد بالكامل</strong>
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### 📚 الدرجات")

    for subject_name, grades in student_data['subjects'].items():
        sa3y = grades['sa3y']
        mid = grades['mid']

        if sa3y is None and mid is None:
            st.markdown(f"""
            <div class="grade-card-missing">
                <div class="subject-name">📖 {subject_name}</div>
                <div class="grade-row">
                    <span class="grade-missing">⏳ لا يوجد درجة لحد الآن</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            sa3y_display = f"{sa3y:g} / 40" if sa3y is not None else '<span class="grade-missing">لا يوجد</span>'
            mid_display = f"{mid:g} / 10" if mid is not None else '<span class="grade-missing">لا يوجد</span>'
            total = (sa3y or 0) + (mid or 0)

            st.markdown(f"""
            <div class="grade-card">
                <div class="subject-name">📖 {subject_name}</div>
                <div class="grade-row">
                    <span class="grade-label">درجة السعي (من 40):</span>
                    <span class="grade-value">{sa3y_display}</span>
                </div>
                <div class="grade-row">
                    <span class="grade-label">درجة الامتحان النصفي (من 10):</span>
                    <span class="grade-value">{mid_display}</span>
                </div>
                <div class="grade-row" style="border-top: 1px solid #e2e8f0; margin-top: 0.5rem; padding-top: 0.7rem;">
                    <span class="grade-label" style="font-weight: 700;">المجموع (من 50):</span>
                    <span class="grade-value" style="background: #c6f6d5; color: #22543d;">{total:g} / 50</span>
                </div>
            </div>
            """, unsafe_allow_html=True)


def display_payment_warning(student_data, remaining, stage_name):
    """عرض تحذير عدم تسديد القسط"""
    remaining_display = f"{remaining:,.0f}" if remaining is not None else "غير محدد"

    st.markdown(f"""
    <div class="student-info">
        <h3>📋 معلومات الطالب</h3>
        <p style="font-size: 1.2rem; margin: 0.5rem 0;">
            <strong>الاسم:</strong> {student_data['name']}
        </p>
        <p style="font-size: 1rem; margin: 0; color: #2d5a8e;">
            <strong>المرحلة:</strong> {stage_name}
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.error(f"""
    ### ⚠️ لا يمكن عرض النتائج

    **يجب دفع القسط المتبقي لعرض درجاتك**

    💰 **المبلغ المتبقي:** {remaining_display} د.ع

    ---

    📌 **ملاحظة مهمة:**
    إذا كان هناك أي خلل في المبلغ المتبقي أو تم التسديد ولم يتم تحديث المعلومات،
    يرجى **مراجعة الشعبة المالية في الجامعة** في أقرب وقت ممكن.
    """)


# ===================== الواجهة الرئيسية =====================
def main():
    st.markdown("""
    <div class="header-box">
        <h1>🎓 نظام عرض نتائج الطلاب</h1>
        <p>قسم هندسة الليزر والإلكترونيات البصرية</p>
        <div class="stage-badge">المرحلتان الثانية والثالثة</div>
    </div>
    """, unsafe_allow_html=True)

    all_data, errors = load_all_stages()

    if errors:
        for err in errors:
            st.error(err)

    if not all_data:
        st.error("❌ لم يتم تحميل أي بيانات. تأكد من وجود الملفات في مجلد التطبيق.")
        return

    st.markdown("### 🔐 الرجاء إدخال الكود الخاص بك")
    st.markdown("<p style='text-align:center; color:#718096; font-size:0.95rem; margin-bottom:1rem;'>الأكواد التي تبدأ بـ <strong>2</strong> للمرحلة الثانية، والتي تبدأ بـ <strong>3</strong> للمرحلة الثالثة</p>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        code_input = st.text_input(
            "الكود",
            placeholder="مثال: 2001 أو 3001",
            label_visibility="collapsed",
            key="student_code"
        )
        check_button = st.button("🔍 عرض النتيجة", use_container_width=True)

    if check_button or code_input:
        if not code_input or not code_input.strip():
            st.warning("⚠️ الرجاء إدخال الكود الخاص بك")
            return

        code = code_input.strip()

        # توحيد الكود (إزالة .0 إن وُجدت)
        try:
            code = str(int(float(code)))
        except (ValueError, TypeError):
            pass

        # تحديد المرحلة من الرقم الأول في الكود
        stage_key = detect_stage(code)

        if stage_key is None:
            st.error(f"""
            ❌ **الكود غير صحيح**

            الكود يجب أن يبدأ بـ **2** (للمرحلة الثانية) أو **3** (للمرحلة الثالثة).

            مثال صحيح: `2001` أو `3015`
            """)
            return

        if stage_key not in all_data:
            st.error(f"❌ بيانات {STAGES[stage_key]['name']} غير متوفرة حالياً.")
            return

        stage_data = all_data[stage_key]
        students = stage_data['students']
        payments = stage_data['payments']
        stage_name = stage_data['name']

        student = students.get(code)

        if student is None:
            st.error(f"""
            ❌ الكود **{code}** غير موجود في {stage_name}.

            الرجاء التأكد من الكود الصحيح أو مراجعة إدارة القسم.
            """)
            return

        remaining, match_type = find_payment_match(
            student['normalized_name'], payments
        )

        if remaining is None:
            if match_type == 'ambiguous':
                st.warning(f"""
                ⚠️ **تم العثور على أكثر من طالب بنفس الاسم في ملف الأقساط**

                الطالب: **{student['name']}** ({stage_name})

                يرجى مراجعة الشعبة المالية في الجامعة لحل التطابق.
                """)
            else:
                st.warning(f"""
                ⚠️ **معلومات الأقساط غير متوفرة لهذا الطالب**

                الطالب: **{student['name']}** ({stage_name})

                يرجى مراجعة الشعبة المالية في الجامعة لتحديث بياناتك.
                """)
            return

        if remaining <= 0:
            display_student_results(student, stage_name)
        else:
            display_payment_warning(student, remaining, stage_name)

    # التذييل
    st.markdown("""
    <div class="footer">
        © 2026 - قسم هندسة الليزر والإلكترونيات البصرية
        <br>
        لأي استفسار يرجى مراجعة الشعبة المالية أو إدارة القسم
        <br><br>
        <span class="developer">تمت برمجة الموقع من قبل م. م. احمد هاشم</span>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
